from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from adp_id_mapper import (
    MasterRecord,
    escape_excel_formula_text,
    extract_last4,
    map_employee_ids,
    normalize_file_number,
    normalize_name,
    resolve_match,
    resolve_match_with_name_fallback,
)


def test_normalization_helpers() -> None:
    assert extract_last4("53513223") == "3223"
    assert extract_last4("ID-1179") == "1179"
    assert extract_last4("12") == ""

    assert normalize_name("  Mary   Ann ") == "mary ann"
    assert normalize_file_number(" 4321 ") == "004321"
    assert normalize_file_number("WRE975803") == "975803"


def test_resolve_match_disambiguates_by_name() -> None:
    candidates = [
        MasterRecord(
            sin_last4="1234",
            first_name="alice",
            last_name="lee",
            file_number="111111",
            source_row=10,
        ),
        MasterRecord(
            sin_last4="1234",
            first_name="bob",
            last_name="lee",
            file_number="222222",
            source_row=11,
        ),
    ]

    resolution = resolve_match(candidates, agency_first_name="bob", agency_last_name="lee")
    assert resolution.file_number == "222222"
    assert resolution.reason is None


def test_resolve_match_falls_back_to_unique_name() -> None:
    name_candidate = MasterRecord(
        sin_last4="8888",
        first_name="sara",
        last_name="khan",
        file_number="777777",
        source_row=20,
    )

    resolution = resolve_match_with_name_fallback(
        sin_candidates=(),
        name_candidates=[name_candidate],
        agency_first_name="sara",
        agency_last_name="khan",
    )

    assert resolution.file_number == "777777"
    assert resolution.reason is None


def test_resolve_match_does_not_override_ambiguous_sin() -> None:
    sin_candidates = [
        MasterRecord(
            sin_last4="1234",
            first_name="alice",
            last_name="lee",
            file_number="111111",
            source_row=10,
        ),
        MasterRecord(
            sin_last4="1234",
            first_name="bob",
            last_name="lee",
            file_number="222222",
            source_row=11,
        ),
    ]
    unique_name_candidate = MasterRecord(
        sin_last4="9999",
        first_name="sara",
        last_name="khan",
        file_number="777777",
        source_row=12,
    )

    resolution = resolve_match_with_name_fallback(
        sin_candidates=sin_candidates,
        name_candidates=[unique_name_candidate],
        agency_first_name="sara",
        agency_last_name="khan",
    )

    assert resolution.file_number is None
    assert resolution.reason is not None
    assert "Ambiguous" in resolution.reason


def test_escape_excel_formula_text() -> None:
    assert escape_excel_formula_text("=SUM(A1:A2)") == "'=SUM(A1:A2)"
    assert escape_excel_formula_text("+abc") == "'+abc"
    assert escape_excel_formula_text("-123") == "'-123"
    assert escape_excel_formula_text("@formula") == "'@formula"
    assert escape_excel_formula_text("safe value") == "safe value"


def test_map_employee_ids_integration(tmp_path: Path) -> None:
    agency_path = tmp_path / "agency.xlsx"
    master_path = tmp_path / "master.xlsx"
    output_path = tmp_path / "agency_adp_mapped.xlsx"

    agency_wb = Workbook()
    agency_ws = agency_wb.active
    agency_ws.title = "P&C"
    agency_ws.append(
        [
            "Employee ID",
            "Employee Last Name",
            "Employee First Name",
            "Hours",
        ]
    )
    agency_ws.append(["7772333", "Smith", "John", 40])
    agency_ws.append(["8881234", "Lee", "Bob", 38])
    agency_ws.append(["9999999", "Doe", "Jane", 36])
    agency_ws.append(["5550000", "Khan", "Sara", 35])
    agency_wb.save(agency_path)

    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "1"
    master_ws.append(["Validation report note row"])
    master_ws.append([None])
    master_ws.append(
        [
            "Employee First Name",
            "Employee Last  Name",
            "File Number",
            "Tax ID (SIN)",
        ]
    )
    master_ws.append(["John", "Smith", "123456", "111 222 333"])
    master_ws.append(["Alice", "Lee", "111111", "111 111 234"])
    master_ws.append(["Bob", "Lee", "222222", "222 111 234"])
    master_ws.append(["Jane", "Doe", "333333", "444 449 999"])
    master_ws.append(["Jane", "Doe", "444444", "777 779 999"])
    master_ws.append(["Sara", "Khan", "777777", "999 668 888"])
    master_wb.save(master_path)

    stats = map_employee_ids(
        agency_workbook_path=agency_path,
        master_workbook_path=master_path,
        output_path=output_path,
        exceptions_sheet_name="Exceptions",
    )

    assert stats.processed_rows == 4
    assert stats.mapped_rows == 3
    assert stats.unchanged_rows == 1
    assert stats.exceptions_count == 1

    out_wb = load_workbook(output_path)
    out_ws = out_wb["P&C"]

    assert out_ws.cell(row=2, column=1).value == "123456"
    assert out_ws.cell(row=3, column=1).value == "222222"
    assert out_ws.cell(row=4, column=1).value == "9999999"
    assert out_ws.cell(row=5, column=1).value == "777777"

    exceptions_ws = out_wb["Exceptions"]
    assert exceptions_ws.max_row == 2
    assert "Ambiguous" in str(exceptions_ws.cell(row=2, column=6).value)


def test_total_hours_sheet_writes_file_number_column(tmp_path: Path) -> None:
    agency_path = tmp_path / "hours.xlsx"
    master_path = tmp_path / "master.xlsx"
    output_path = tmp_path / "hours_adp_mapped.xlsx"

    agency_wb = Workbook()
    agency_hours_ws = agency_wb.active
    agency_hours_ws.title = "Agency Hours"
    agency_hours_ws.append(
        [
            "Employee ID",
            "Employee Last Name",
            "Employee First Name",
            "Hours",
        ]
    )
    agency_hours_ws.append(["9990000", "Ignore", "Me", 10])

    total_hours_ws = agency_wb.create_sheet("Total Hours")
    total_hours_ws.append(
        [
            "File # ",
            "Employee ID",
            "Employee Last Name",
            "Employee First Name",
            "Total Regular  Hours ",
        ]
    )
    total_hours_ws.append([None, "5355118", "Aery", "Meenu", 65.1])
    agency_wb.save(agency_path)

    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "1"
    master_ws.append(
        [
            "Employee First Name",
            "Employee Last  Name",
            "File Number",
            "Tax ID (SIN)",
        ]
    )
    master_ws.append(["Meenu", "Aery", "975806", "111 111 118"])
    master_wb.save(master_path)

    stats = map_employee_ids(
        agency_workbook_path=agency_path,
        master_workbook_path=master_path,
        output_path=output_path,
        exceptions_sheet_name="Exceptions",
        agency_sheet_name="Total Hours",
    )

    assert stats.processed_rows == 1
    assert stats.mapped_rows == 1
    assert stats.unchanged_rows == 0
    assert stats.exceptions_count == 0

    out_wb = load_workbook(output_path)
    out_total_hours_ws = out_wb["Total Hours"]
    out_agency_hours_ws = out_wb["Agency Hours"]

    # Auto target column prefers File # when available.
    assert out_total_hours_ws.cell(row=2, column=1).value == "975806"
    # Source Employee ID remains unchanged in Total Hours sheet.
    assert out_total_hours_ws.cell(row=2, column=2).value == "5355118"
    # Non-target sheets are untouched.
    assert out_agency_hours_ws.cell(row=2, column=1).value == "9990000"


def test_file_sharp_header_auto_target(tmp_path: Path) -> None:
    agency_path = tmp_path / "hours_file_sharp.xlsx"
    master_path = tmp_path / "master_file_sharp.xlsx"
    output_path = tmp_path / "hours_file_sharp_out.xlsx"

    agency_wb = Workbook()
    ws = agency_wb.active
    ws.title = "Total Hours"
    ws.append(["File#", "Employee ID", "Employee Last Name", "Employee First Name"])
    ws.append([None, "5355118", "Aery", "Meenu"])
    agency_wb.save(agency_path)

    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "1"
    master_ws.append(
        [
            "Employee First Name",
            "Employee Last  Name",
            "File Number",
            "Tax ID (SIN)",
        ]
    )
    master_ws.append(["Meenu", "Aery", "975806", "999 995 118"])
    master_wb.save(master_path)

    map_employee_ids(
        agency_workbook_path=agency_path,
        master_workbook_path=master_path,
        output_path=output_path,
        exceptions_sheet_name="Exceptions",
        agency_sheet_name="Total Hours",
    )

    out_wb = load_workbook(output_path)
    out_ws = out_wb["Total Hours"]

    assert out_ws.cell(row=2, column=1).value == "975806"
    assert out_ws.cell(row=2, column=2).value == "5355118"
