from __future__ import annotations

import argparse
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LOGGER = logging.getLogger("adp_id_mapper")

HEADER_SCAN_ROW_LIMIT = 60
HEADER_SCAN_COLUMN_LIMIT = 200

AGENCY_REQUIRED_HEADERS = frozenset(
    {
        "employee id",
        "employee first name",
        "employee last name",
    }
)
MASTER_REQUIRED_HEADERS = frozenset(
    {
        "tax id (sin)",
        "employee first name",
        "employee last name",
        "file number",
    }
)


@dataclass(frozen=True)
class HeaderInfo:
    row_index: int
    column_by_header: Mapping[str, int]


@dataclass(frozen=True)
class MasterRecord:
    sin_last4: str
    first_name: str
    last_name: str
    file_number: str
    source_row: int


@dataclass(frozen=True)
class MatchResolution:
    file_number: str | None
    reason: str | None
    candidate_file_numbers: tuple[str, ...]


@dataclass(frozen=True)
class MappingStats:
    processed_rows: int
    mapped_rows: int
    unchanged_rows: int
    exceptions_count: int
    output_path: Path


def canonical_header(value: object) -> str:
    if value is None:
        return ""
    header = str(value).strip().lower()
    return re.sub(r"\s+", " ", header)


def digits_only(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value) if ch.isdigit())


def extract_last4(value: object) -> str:
    digits = digits_only(value)
    if len(digits) < 4:
        return ""
    return digits[-4:]


def normalize_name(value: object) -> str:
    if value is None:
        return ""
    lowered = str(value).strip().lower()
    return re.sub(r"\s+", " ", lowered)


def normalize_file_number(value: object) -> str:
    digits = digits_only(value)
    if not digits:
        return ""
    if len(digits) > 6:
        return ""
    return digits.zfill(6)


def escape_excel_formula_text(value: object) -> str:
    text = str(value or "")
    if text and text[0] in ("=", "+", "-", "@"):
        return f"'{text}"
    return text


def find_header_info(
    worksheet: Worksheet,
    required_headers: frozenset[str],
) -> HeaderInfo | None:
    max_row = min(worksheet.max_row, HEADER_SCAN_ROW_LIMIT)
    max_col = min(worksheet.max_column, HEADER_SCAN_COLUMN_LIMIT)

    for row_idx in range(1, max_row + 1):
        column_by_header: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            key = canonical_header(value)
            if key:
                column_by_header[key] = col_idx

        if required_headers.issubset(column_by_header.keys()):
            return HeaderInfo(row_index=row_idx, column_by_header=column_by_header)

    return None


def find_sheet_and_header(
    workbook: Workbook,
    required_headers: frozenset[str],
    workbook_label: str,
) -> tuple[Worksheet, HeaderInfo]:
    for worksheet in workbook.worksheets:
        header_info = find_header_info(worksheet, required_headers)
        if header_info is not None:
            return worksheet, header_info

    missing = ", ".join(sorted(required_headers))
    raise ValueError(
        f"Could not find a sheet in {workbook_label} with required headers: {missing}"
    )


def build_master_index(
    worksheet: Worksheet,
    header_info: HeaderInfo,
) -> dict[str, list[MasterRecord]]:
    first_col = header_info.column_by_header["employee first name"]
    last_col = header_info.column_by_header["employee last name"]
    sin_col = header_info.column_by_header["tax id (sin)"]
    file_col = header_info.column_by_header["file number"]

    index: dict[str, list[MasterRecord]] = {}

    for row_idx in range(header_info.row_index + 1, worksheet.max_row + 1):
        first_name = normalize_name(worksheet.cell(row=row_idx, column=first_col).value)
        last_name = normalize_name(worksheet.cell(row=row_idx, column=last_col).value)
        sin_last4 = extract_last4(worksheet.cell(row=row_idx, column=sin_col).value)
        file_number = normalize_file_number(worksheet.cell(row=row_idx, column=file_col).value)

        if not first_name or not last_name or not sin_last4 or not file_number:
            continue

        record = MasterRecord(
            sin_last4=sin_last4,
            first_name=first_name,
            last_name=last_name,
            file_number=file_number,
            source_row=row_idx,
        )
        index.setdefault(sin_last4, []).append(record)

    return index


def build_name_index(
    master_index: Mapping[str, Sequence[MasterRecord]],
) -> dict[tuple[str, str], list[MasterRecord]]:
    name_index: dict[tuple[str, str], list[MasterRecord]] = {}
    for records in master_index.values():
        for record in records:
            key = (record.first_name, record.last_name)
            name_index.setdefault(key, []).append(record)
    return name_index


def resolve_match(
    candidates: Sequence[MasterRecord],
    agency_first_name: str,
    agency_last_name: str,
) -> MatchResolution:
    if not candidates:
        return MatchResolution(
            file_number=None,
            reason="No master record has matching SIN last-4 digits.",
            candidate_file_numbers=(),
        )

    if len(candidates) == 1:
        return MatchResolution(
            file_number=candidates[0].file_number,
            reason=None,
            candidate_file_numbers=(candidates[0].file_number,),
        )

    name_matches = [
        candidate
        for candidate in candidates
        if candidate.first_name == agency_first_name
        and candidate.last_name == agency_last_name
    ]

    if len(name_matches) == 1:
        return MatchResolution(
            file_number=name_matches[0].file_number,
            reason=None,
            candidate_file_numbers=(name_matches[0].file_number,),
        )

    candidate_ids = tuple(sorted({candidate.file_number for candidate in candidates}))
    if len(name_matches) > 1:
        reason = (
            "Ambiguous: multiple master records share SIN last-4 and normalized name."
        )
    else:
        reason = (
            "Ambiguous: multiple master records share SIN last-4 and no unique name match."
        )

    return MatchResolution(
        file_number=None,
        reason=reason,
        candidate_file_numbers=candidate_ids,
    )


def resolve_match_with_name_fallback(
    sin_candidates: Sequence[MasterRecord],
    name_candidates: Sequence[MasterRecord],
    agency_first_name: str,
    agency_last_name: str,
) -> MatchResolution:
    sin_resolution = resolve_match(
        sin_candidates,
        agency_first_name=agency_first_name,
        agency_last_name=agency_last_name,
    )
    if sin_resolution.file_number is not None:
        return sin_resolution

    # If SIN candidates exist but are ambiguous, do not override with name fallback.
    if sin_candidates:
        return sin_resolution

    if len(name_candidates) == 1:
        fallback = name_candidates[0]
        return MatchResolution(
            file_number=fallback.file_number,
            reason=None,
            candidate_file_numbers=(fallback.file_number,),
        )

    if len(name_candidates) > 1:
        return MatchResolution(
            file_number=None,
            reason=(
                "Ambiguous: multiple master records share normalized first and last name."
            ),
            candidate_file_numbers=tuple(
                sorted({candidate.file_number for candidate in name_candidates})
            ),
        )

    return sin_resolution


def default_output_path(agency_workbook_path: Path) -> Path:
    return agency_workbook_path.with_name(
        f"{agency_workbook_path.stem}_adp_mapped{agency_workbook_path.suffix}"
    )


def write_exceptions_sheet(
    workbook: Workbook,
    exceptions: Sequence[dict[str, str | int]],
    exceptions_sheet_name: str,
) -> None:
    if exceptions_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[exceptions_sheet_name])

    exceptions_sheet = workbook.create_sheet(exceptions_sheet_name)
    exceptions_sheet.append(
        [
            "Agency Row",
            "Original Employee ID",
            "Employee Last Name",
            "Employee First Name",
            "SIN Last 4 Used",
            "Reason",
            "Candidate File Numbers",
        ]
    )

    for item in exceptions:
        exceptions_sheet.append(
            [
                item["agency_row"],
                escape_excel_formula_text(item["original_employee_id"]),
                escape_excel_formula_text(item["employee_last_name"]),
                escape_excel_formula_text(item["employee_first_name"]),
                escape_excel_formula_text(item["sin_last4"]),
                escape_excel_formula_text(item["reason"]),
                escape_excel_formula_text(item["candidate_file_numbers"]),
            ]
        )

    exceptions_sheet.freeze_panes = "A2"


def map_employee_ids(
    agency_workbook_path: Path,
    master_workbook_path: Path,
    output_path: Path | None,
    exceptions_sheet_name: str,
) -> MappingStats:
    agency_workbook = load_workbook(agency_workbook_path)
    master_workbook = load_workbook(master_workbook_path, data_only=True)

    try:
        agency_sheet, agency_header = find_sheet_and_header(
            agency_workbook,
            AGENCY_REQUIRED_HEADERS,
            workbook_label=str(agency_workbook_path),
        )
        master_sheet, master_header = find_sheet_and_header(
            master_workbook,
            MASTER_REQUIRED_HEADERS,
            workbook_label=str(master_workbook_path),
        )

        master_index = build_master_index(master_sheet, master_header)
        master_name_index = build_name_index(master_index)

        id_col = agency_header.column_by_header["employee id"]
        first_col = agency_header.column_by_header["employee first name"]
        last_col = agency_header.column_by_header["employee last name"]

        processed_rows = 0
        mapped_rows = 0
        unchanged_rows = 0
        exceptions: list[dict[str, str | int]] = []

        for row_idx in range(agency_header.row_index + 1, agency_sheet.max_row + 1):
            original_id = agency_sheet.cell(row=row_idx, column=id_col).value
            first_name_raw = agency_sheet.cell(row=row_idx, column=first_col).value
            last_name_raw = agency_sheet.cell(row=row_idx, column=last_col).value

            if not any(
                str(value).strip() if value is not None else ""
                for value in (original_id, first_name_raw, last_name_raw)
            ):
                continue

            processed_rows += 1

            first_name = normalize_name(first_name_raw)
            last_name = normalize_name(last_name_raw)
            sin_last4 = extract_last4(original_id)

            if not sin_last4:
                sin_candidates: Sequence[MasterRecord] = ()
            else:
                sin_candidates = master_index.get(sin_last4, [])

            name_candidates = master_name_index.get((first_name, last_name), [])
            resolution = resolve_match_with_name_fallback(
                sin_candidates,
                name_candidates,
                agency_first_name=first_name,
                agency_last_name=last_name,
            )

            if resolution.file_number is not None:
                target_cell = agency_sheet.cell(row=row_idx, column=id_col)
                target_cell.value = resolution.file_number
                target_cell.number_format = "@"
                mapped_rows += 1
            else:
                unchanged_rows += 1
                reason = resolution.reason or "No match."
                if not sin_last4:
                    if len(name_candidates) > 1:
                        reason = (
                            "Employee ID does not contain at least 4 digits and name match was not unique."
                        )
                    elif len(name_candidates) == 0:
                        reason = (
                            "Employee ID does not contain at least 4 digits and no name match was found."
                        )
                exceptions.append(
                    {
                        "agency_row": row_idx,
                        "original_employee_id": str(original_id or ""),
                        "employee_last_name": str(last_name_raw or ""),
                        "employee_first_name": str(first_name_raw or ""),
                        "sin_last4": sin_last4,
                        "reason": reason,
                        "candidate_file_numbers": ", ".join(
                            resolution.candidate_file_numbers
                        ),
                    }
                )

        write_exceptions_sheet(
            agency_workbook,
            exceptions,
            exceptions_sheet_name=exceptions_sheet_name,
        )

        final_output_path = output_path or default_output_path(agency_workbook_path)
        final_output_path.parent.mkdir(parents=True, exist_ok=True)
        agency_workbook.save(final_output_path)

        return MappingStats(
            processed_rows=processed_rows,
            mapped_rows=mapped_rows,
            unchanged_rows=unchanged_rows,
            exceptions_count=len(exceptions),
            output_path=final_output_path,
        )
    finally:
        agency_workbook.close()
        master_workbook.close()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Replace Employee ID values in an Agency workbook with ADP File Numbers "
            "using a master validation workbook."
        )
    )
    parser.add_argument("--agency", required=True, help="Path to Agency workbook (.xlsx)")
    parser.add_argument("--master", required=True, help="Path to master workbook (.xlsx)")
    parser.add_argument(
        "--output",
        required=False,
        help=(
            "Output workbook path. If omitted, writes beside agency file with "
            "_adp_mapped suffix."
        ),
    )
    parser.add_argument(
        "--exceptions-sheet-name",
        default="Exceptions",
        help="Worksheet name used for unresolved/ambiguous rows.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable debug logging.",
    )
    return parser.parse_args(argv)


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    configure_logging(verbose=args.verbose)

    agency_path = Path(args.agency).expanduser().resolve()
    master_path = Path(args.master).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else None

    if not agency_path.exists():
        LOGGER.error("Agency workbook was not found: %s", agency_path)
        return 2
    if not master_path.exists():
        LOGGER.error("Master workbook was not found: %s", master_path)
        return 2

    try:
        stats = map_employee_ids(
            agency_workbook_path=agency_path,
            master_workbook_path=master_path,
            output_path=output_path,
            exceptions_sheet_name=args.exceptions_sheet_name,
        )
    except Exception:
        LOGGER.exception("Failed while mapping employee IDs.")
        return 1

    LOGGER.info("Processed rows: %s", stats.processed_rows)
    LOGGER.info("Mapped rows: %s", stats.mapped_rows)
    LOGGER.info("Unchanged rows: %s", stats.unchanged_rows)
    LOGGER.info("Exception rows: %s", stats.exceptions_count)
    LOGGER.info("Output workbook: %s", stats.output_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
